#!/usr/bin/env python3
"""
原始数据自动抓取器 — 从Nature/BMC/PMC自动下载Supplementary Excel/CSV

支持的来源:
- Nature系列 (nature.com): 2022年后强制上传原始数据
- BMC系列 (biomedcentral.com): 补充材料页面
- PubMed Central (ncbi.nlm.nih.gov/pmc): 补充材料入口
"""

import json
import os
import re
import sys
import tempfile
import zipfile
from pathlib import Path
from urllib.parse import urljoin, urlparse

try:
    import requests
    from bs4 import BeautifulSoup
except ImportError:
    print("需要: pip install requests beautifulsoup4")
    sys.exit(1)

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36'
}


class SupplementaryDataFetcher:
    """从论文DOI/URL自动下载原始数据"""

    def __init__(self, output_dir: str = None):
        self.output_dir = Path(output_dir) if output_dir else Path(tempfile.mkdtemp())
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.downloaded = []
        self.tables = []

    def fetch(self, doi_or_pmcid: str) -> dict:
        """主入口: 给定DOI或PMCID，下载并解析所有Supplementary Excel/CSV"""
        # 判断输入类型
        if doi_or_pmcid.startswith('PMC'):
            return self._fetch_from_pmc(doi_or_pmcid)
        else:
            return self._fetch_by_doi(doi_or_pmcid)

    def _fetch_by_doi(self, doi: str) -> dict:
        """通过DOI获取论文页面，定位补充材料"""
        doi = doi.strip().replace('https://doi.org/', '')

        # 尝试Nature/BMC入口
        url = f"https://doi.org/{doi}"
        try:
            resp = requests.get(url, headers=HEADERS, timeout=30, allow_redirects=True)
            final_url = resp.url
        except Exception as e:
            return {"error": f"DOI resolution failed: {e}", "tables": [], "downloaded_files": []}

        html = resp.text
        soup = BeautifulSoup(html, 'html.parser')

        excel_links = []

        # Pattern 1: Nature系列 — "Source Data" / "Supplementary Data" 链接
        for a in soup.find_all('a', href=True):
            href = a['href']
            text = (a.get_text() or '').lower()
            if any(kw in text for kw in ['source data', 'supplementary data', 'supplementary information']):
                full_url = urljoin(final_url, href)
                excel_links.append(full_url)

        # Pattern 2: 直接搜索 .xlsx / .csv 链接
        for a in soup.find_all('a', href=True):
            href = a['href']
            if re.search(r'\.(xlsx|csv)$', href, re.I):
                full_url = urljoin(final_url, href)
                if full_url not in excel_links:
                    excel_links.append(full_url)

        # Pattern 3: Nature 标准化 Supplementary Information 页面
        # 通常是 https://www.nature.com/articles/XXXXX#SecXX
        if not excel_links:
            article_id = None
            m = re.search(r'/articles/([^/#]+)', final_url)
            if m:
                article_id = m.group(1)
                si_url = f"https://www.nature.com/articles/{article_id}#Sec"
                try:
                    si_resp = requests.get(si_url, headers=HEADERS, timeout=30)
                    si_soup = BeautifulSoup(si_resp.text, 'html.parser')
                    for a in si_soup.find_all('a', href=True):
                        if re.search(r'\.(xlsx|csv|xls)$', a['href'], re.I):
                            full_url = urljoin(si_url, a['href'])
                            if full_url not in excel_links:
                                excel_links.append(full_url)
                except Exception:
                    pass

        if not excel_links:
            return {
                "doi": doi,
                "error": "No supplementary Excel/CSV files found",
                "tables": [],
                "downloaded_files": []
            }

        # 下载文件
        for url in excel_links:
            try:
                fname = Path(urlparse(url).path).name
                if not re.search(r'\.(xlsx|xls|csv|zip)$', fname, re.I):
                    fname += '.xlsx'

                filepath = self.output_dir / fname
                dl_resp = requests.get(url, headers=HEADERS, timeout=60)
                dl_resp.raise_for_status()

                with open(filepath, 'wb') as f:
                    f.write(dl_resp.content)
                self.downloaded.append(str(filepath))

                # 解析
                self._parse_file(filepath)

            except Exception as e:
                print(f"  ⚠ 下载失败 {url}: {e}", file=sys.stderr)

        # 按sheet名分组
        tables_by_source = {}
        for t in self.tables:
            bucket = t.get('source_file', 'unknown')
            if bucket not in tables_by_source:
                tables_by_source[bucket] = []
            tables_by_source[bucket].append(t)

        return {
            "doi": doi,
            "tables": self.tables,
            "tables_by_source": tables_by_source,
            "downloaded_files": self.downloaded,
            "total_tables": len(self.tables)
        }

    def _fetch_from_pmc(self, pmcid: str) -> dict:
        """从PubMed Central获取补充材料"""
        pmcid = pmcid.strip()
        # PMC补充材料通常是 https://www.ncbi.nlm.nih.gov/pmc/articles/{PMCID}/
        url = f"https://www.ncbi.nlm.nih.gov/pmc/articles/{pmcid}/"
        try:
            resp = requests.get(url, headers=HEADERS, timeout=30)
        except Exception as e:
            return {"error": f"PMC access failed: {e}", "tables": [], "downloaded_files": []}

        soup = BeautifulSoup(resp.text, 'html.parser')
        excel_links = []

        for a in soup.find_all('a', href=True):
            href = a['href']
            if re.search(r'\.(xlsx|xls|csv)$', href, re.I):
                full_url = urljoin(url, href)
                excel_links.append(full_url)

        for url in excel_links:
            try:
                fname = Path(urlparse(url).path).name
                filepath = self.output_dir / fname
                dl_resp = requests.get(url, headers=HEADERS, timeout=60)
                dl_resp.raise_for_status()
                with open(filepath, 'wb') as f:
                    f.write(dl_resp.content)
                self.downloaded.append(str(filepath))
                self._parse_file(filepath)
            except Exception as e:
                print(f"  ⚠ PMC下载失败: {e}", file=sys.stderr)

        return {
            "pmcid": pmcid,
            "tables": self.tables,
            "downloaded_files": self.downloaded,
            "total_tables": len(self.tables)
        }

    def _parse_file(self, filepath: str):
        """解析下载的Excel/CSV，提取所有数值表格"""
        path = Path(filepath)

        if path.suffix.lower() in ('.xlsx', '.xls') and HAS_OPENPYXL:
            try:
                wb = openpyxl.load_workbook(filepath, data_only=True)
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    rows = []
                    headers = []
                    for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                        if row_idx == 0:
                            headers = [str(c) if c else '' for c in row]
                        else:
                            rows.append(list(row))
                    if rows and self._has_numeric_data(rows):
                        self.tables.append({
                            "source_file": path.name,
                            "sheet": sheet_name,
                            "headers": headers,
                            "rows": rows[:5000],  # 限制防止爆炸
                            "row_count": len(rows)
                        })
            except Exception as e:
                print(f"  ⚠ 解析Excel失败 {path.name}: {e}", file=sys.stderr)

        elif path.suffix.lower() == '.csv':
            import csv
            try:
                with open(filepath, 'r', encoding='utf-8-sig') as f:
                    reader = csv.reader(f)
                    headers = next(reader, [])
                    rows = [list(row) for row in reader]
                if rows and self._has_numeric_data(rows):
                    self.tables.append({
                        "source_file": path.name,
                        "sheet": "Sheet1",
                        "headers": [str(h) for h in headers],
                        "rows": rows[:5000],
                        "row_count": len(rows)
                    })
            except Exception as e:
                print(f"  ⚠ 解析CSV失败: {e}", file=sys.stderr)

        elif path.suffix.lower() == '.zip':
            try:
                with zipfile.ZipFile(filepath) as zf:
                    for name in zf.namelist():
                        if re.search(r'\.(xlsx|xls|csv)$', name, re.I):
                            extracted = zf.extract(name, self.output_dir)
                            self._parse_file(extracted)
            except Exception as e:
                print(f"  ⚠ 解压ZIP失败: {e}", file=sys.stderr)

    @staticmethod
    def _has_numeric_data(rows: list, threshold: float = 0.3) -> bool:
        """检查表格是否包含足够多的数值数据"""
        numeric_count = 0
        total = 0
        for row in rows:
            for cell in row:
                total += 1
                if isinstance(cell, (int, float)):
                    numeric_count += 1
                elif isinstance(cell, str):
                    try:
                        float(cell.replace('%', '').replace(',', ''))
                        numeric_count += 1
                    except ValueError:
                        pass
        return total > 0 and numeric_count / total > threshold


def fetch_supplementary_data(doi_or_pmcid: str, output_dir: str = None) -> dict:
    """便捷函数: 给定DOI/PMCID，下载并返回数值表格"""
    fetcher = SupplementaryDataFetcher(output_dir)
    return fetcher.fetch(doi_or_pmcid)


def main():
    import argparse
    parser = argparse.ArgumentParser(description='原始数据自动抓取器')
    parser.add_argument('identifier', help='DOI或PMCID')
    parser.add_argument('-o', '--output-dir', help='输出目录')
    parser.add_argument('--dry-run', action='store_true', help='仅列出可下载文件，不实际下载')
    args = parser.parse_args()

    fetcher = SupplementaryDataFetcher(args.output_dir)
    if args.dry_run:
        # 仅探测
        result = fetcher._fetch_by_doi(args.identifier)
        result.pop('tables', None)
        result.pop('rows', None)
    else:
        result = fetcher.fetch(args.identifier)

    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == '__main__':
    main()
