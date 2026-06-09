#!/usr/bin/env python3
"""
等差数列/固定比值自动检测器 — 学术不端数学检测的自动化版本 v6.4

检测论文数据中的数学规律异常：
1. 任意两列间的固定差值 / 固定比值 / 线性关系
2. 单列内的等差数列
3. 末位数字集中度（0+5占比，全是某数字）
4. 小数点后位数异常
5. 块级结构重复
6. byte-identical重复 (公开方法方法论: 逐字节相同的行/块/列)
7. 🆕 v6.0 常数平移检测 (已知公开案例: D3=GlueBody2+0.02)
8. 🆕 v6.0 跨文件/跨图固定差值配对 (已知公开案例: 图3G vs 扩展图7B 差41.4)
9. 🆕 v6.0 行级复制粘贴检测 (已知公开案例: 实验组前三行=对照组前三行)

输出: JSON格式，每个发现含检测类型、规律描述、数学不可能概率

🆕 v6.0 已知公开案例技术落地 (2026-06-06):
- 新增检测10: 常数平移检测 (CONSTANT_SHIFT) — 一个序列是另一个序列的常数平移
  区别现有检测1(固定差值): 检测1看"多个行差值相同"，常数平移看"整个序列差值是固定常数"
  检测1需要大量行(>=3且>=50%行差值相同)才触发，3个点的平移会被遗漏。
  案例: GlueBody2 [1.03,0.99,0.97] → D3 [1.05,1.01,0.99]，对应位差均为+0.02
- 新增检测11: 跨文件/跨sheet固定差值配对 (CROSS_FILE_FIXED_DIFFERENCE)
  从多个CSV/sheet加载所有数值列 → 跨文件两两配对 → 检测固定差值
  区别于单文件内检测: 造假者把数据分散到不同文件/sheet避免被单文件检测捕获
  案例: 图3G [33.8,33.9,37.9] 与 扩展图7B [75.2,75.3,79.3]，每对应位差=41.4
- 新增检测12: 跨组行级复制粘贴 (CROSS_GROUP_ROW_DUPLICATION)
  两组声称独立的实验数据，前N行逐行完全相同但后续不同
  区别于检测7(块级重复): 检测7看"整表内任意重复块"，检测12专看"两组的行逐行相同"
  案例: GlueBody前三个样本 [14.5,16.0,21.5] = GlueBody2前三个样本 [14.5,16.0,21.5]

⚠️ 已知盲区 (v5.0 实测):
- 固定差值/等差数列检测对带噪声的编造数据失效。当造假者在每个数据点加入
  ±0.2-0.3的随机抖动后，精确数学规律被稀释，固定差值检测返回0异常。
- 末位偏好是最后一个数据层面防线——编造者从"好看"整数(100,80,60...)开始
  加噪声后，末位的0/5偏好仍残留（实测30样本编造=56.7%，真随机=0异常）。
- 这意味着: 末位偏好不会产生假阳性——没信号就是没信号。但固定差值没信号
  不代表数据真实（造假者可能比较"精细"）。
- 实测案例验证: 15篇论文中无表格编造模式（固定差值/等差数列/全末位5均为0），
  但图片取证维度确认L4级造假。图片造假可完全独立于表格数据编造。

v5.2 公开方法方法论增强 (2026-06-02):
- byte-identical重复检测 ✅
  - 行级: 整行所有数值完全逐字节相同
  - 块级: 连续3/4/5行序列在论文中不同位置重复出现
  - 跨列: 两列完全逐值相同
- 公开方法实测: 1446篇Nature论文 → 141篇有source data → AI扫出7篇证据硬的

🆕 v6.4 公开方法方法论盲区补完 (2026-06-06):
- _detect_byte_identical() 策略4: 列复制改1格 (BYTE_IDENTICAL_NEAR_COLUMNS) ✅
  检测两列仅1-2个值不同但其余完全逐字节相同——编辑距离1-2的近似复制
  → 典型的"复制后改1格来避免被发现完全一模一样"
- _detect_arithmetic_grid(): 等差网格检测 (ARITHMETIC_GRID) ✅
  检测唯一值集合本身呈等间距分布——如17个唯一值恰好是等差数列
  → 取值集合内部的等间距规律，比索引间等差更隐蔽
v5.2 已知盲区（公开方法方法论的两种模式已全部补完✅）
"""

import csv
import json
import math
import re
import sys
from collections import Counter
from itertools import combinations
from pathlib import Path

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class ArithmeticSequenceDetector:
    """扫描所有数值列，检测数学规律异常 v6.0"""

    def __init__(self):
        self.findings = []
        self.total_cells = 0
        self.total_columns = 0
        # 🆕 v6.0: 跨文件检测模式
        self.multi_source_mode = False
        self.all_source_columns = {}  # {source_name: {col_name: [values]}}

    def detect_from_file(self, filepath: str) -> list[dict]:
        """从CSV/Excel文件加载数据并检测"""
        path = Path(filepath)
        if path.suffix.lower() in ('.xlsx', '.xls'):
            return self.detect_from_excel(filepath)
        else:
            return self.detect_from_csv(filepath)

    def detect_from_csv(self, filepath: str) -> list[dict]:
        rows = []
        with open(filepath, 'r', encoding='utf-8-sig') as f:
            reader = csv.reader(f)
            headers = next(reader, None)
            for row in reader:
                rows.append(row)
        return self.detect(headers, rows)

    def detect_from_excel(self, filepath: str) -> list[dict]:
        if not HAS_OPENPYXL:
            return [{"error": "openpyxl not installed"}]
        wb = openpyxl.load_workbook(filepath, data_only=True)
        all_findings = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = [[cell.value for cell in row] for row in ws.iter_rows()]
            if not rows:
                continue
            headers = rows[0]
            data_rows = rows[1:]
            findings = self.detect(headers, data_rows)
            for f in findings:
                f['sheet'] = sheet_name
            all_findings.extend(findings)
        return all_findings

    def detect(self, headers: list, rows: list) -> list[dict]:
        """主检测入口"""
        self.findings = []

        # 转换为列式存储
        columns = {}
        for i, h in enumerate(headers):
            col_data = []
            for row in rows:
                if i < len(row):
                    val = self._parse_number(row[i])
                    if val is not None:
                        col_data.append(val)
            if len(col_data) >= 3:
                name = str(h).strip() if h else f"Col_{i}"
                columns[name] = col_data

        if len(columns) < 1:
            return self.findings

        self.total_columns = len(columns)
        self.total_cells = sum(len(v) for v in columns.values())

        # === 检测流水线 ===
        # 单列检测（不需要2列也能跑）
        self._detect_single_column_arithmetic(columns)
        self._detect_digit_preference(columns)
        self._detect_decimal_anomaly(columns, rows, headers)
        self._detect_decimal_prefix_repeat(columns)  # 🆕 v8.4: Decimal前缀重复（公开方法论）
        self._detect_arithmetic_grid(columns)  # 🆕 v6.4: 等差网格（单列也能跑）

        # 多列检测（需要至少2列）
        if len(columns) >= 2:
            self._detect_fixed_difference(columns)
            self._detect_fixed_ratio(columns)
            self._detect_linear_relation(columns)
            self._detect_block_repetition(columns)
            self._detect_noise_pattern(columns)  # 🆕 v5.0: 噪声环境下的编造检测
            self._detect_byte_identical(columns)  # 🆕 v5.2: byte-identical重复检测（公开方法方法论）
            self._detect_constant_shift(columns)  # 🆕 v6.0: 常数平移检测（已知公开案例）
            self._detect_cross_group_row_dup(columns)  # 🆕 v6.0: 跨组行级复制粘贴（已知公开案例）

        return self.findings

    # ==================================================================
    # 检测 1: 固定差值
    # ==================================================================
    def _detect_fixed_difference(self, columns: dict):
        col_names = list(columns.keys())
        for c1, c2 in combinations(col_names, 2):
            diffs = []
            hit_count = 0
            total = min(len(columns[c1]), len(columns[c2]))
            if total < 3:
                continue
            for i in range(total):
                d = round(columns[c2][i] - columns[c1][i], 10)
                diffs.append(d)
            if not diffs:
                continue
            from collections import Counter as C
            diff_counter = C(diffs)
            most_common_diff, freq = diff_counter.most_common(1)[0]

            if freq >= total * 0.5 and freq >= 3 and most_common_diff != 0:
                # 计算数学不可能概率
                # 如果两列独立随机，观测到freq个相同的固定差概率近乎0
                p = self._binomial_prob(freq, total, 0.1)  # 保守先验每个差值概率10%
                self._add_finding(
                    "FIXED_DIFFERENCE",
                    f"列 '{c1}' 和 '{c2}' 固定差值 = {most_common_diff}，共 {freq}/{total} 行 ({freq/total*100:.1f}%)",
                    p,
                    {
                        "column_a": c1, "column_b": c2,
                        "fixed_difference": most_common_diff,
                        "matching_rows": freq, "total_rows": total,
                        "match_rate": freq/total
                    }
                )

    # ==================================================================
    # 检测 2: 固定比值
    # ==================================================================
    def _detect_fixed_ratio(self, columns: dict):
        col_names = list(columns.keys())
        for c1, c2 in combinations(col_names, 2):
            ratios = []
            total = min(len(columns[c1]), len(columns[c2]))
            if total < 3:
                continue
            for i in range(total):
                if abs(columns[c1][i]) > 1e-10:
                    r = round(columns[c2][i] / columns[c1][i], 6)
                    ratios.append(r)
            if not ratios:
                continue
            from collections import Counter as C
            ratio_counter = C(ratios)
            most_common_ratio, freq = ratio_counter.most_common(1)[0]

            if freq >= total * 0.5 and freq >= 3 and abs(most_common_ratio - 1.0) > 1e-6:
                p = self._binomial_prob(freq, total, 0.1)
                self._add_finding(
                    "FIXED_RATIO",
                    f"列 '{c1}' 和 '{c2}' 固定比值 = {most_common_ratio}，共 {freq}/{total} 行 ({freq/total*100:.1f}%)",
                    p,
                    {
                        "column_a": c1, "column_b": c2,
                        "fixed_ratio": most_common_ratio,
                        "matching_rows": freq, "total_rows": total,
                        "match_rate": freq/total
                    }
                )

    # ==================================================================
    # 检测 3: 线性关系
    # ==================================================================
    def _detect_linear_relation(self, columns: dict):
        col_names = list(columns.keys())
        for c1, c2 in combinations(col_names, 2):
            total = min(len(columns[c1]), len(columns[c2]))
            if total < 4:
                continue

            # 检测 a*c1 + b = c2 (整数系数)
            import itertools
            for shift in [-10, -5, -2, -1, 0, 1, 2, 5, 10]:
                for mult in [-5, -2, -1, 1, 2, 5]:
                    matches = 0
                    for i in range(total):
                        predicted = mult * columns[c1][i] + shift
                        if abs(round(predicted, 8) - round(columns[c2][i], 8)) < 1e-6:
                            matches += 1
                    if matches >= total * 0.8 and matches >= 4:
                        p = self._binomial_prob(matches, total, 0.05)
                        desc = f"{c2} = {mult}×{c1}"
                        if shift > 0:
                            desc += f" + {shift}"
                        elif shift < 0:
                            desc += f" - {abs(shift)}"
                        self._add_finding(
                            "LINEAR_RELATION",
                            f"列 '{c1}' 和 '{c2}' 存在线性关系: {desc}，{matches}/{total} 行匹配 ({matches/total*100:.1f}%)",
                            p,
                            {
                                "column_a": c1, "column_b": c2,
                                "multiplier": mult, "shift": shift,
                                "matching_rows": matches, "total_rows": total,
                                "match_rate": matches/total
                            }
                        )

    # ==================================================================
    # 检测 4: 单列等差数列 (如已知案例：70个数据全是等差数列)
    # ==================================================================
    def _detect_single_column_arithmetic(self, columns: dict):
        col_names = list(columns.keys())
        for c in col_names:
            vals = columns[c]
            if len(vals) < 4:
                continue

            # 相邻差值
            diffs = [round(vals[i+1] - vals[i], 10) for i in range(len(vals)-1)]
            from collections import Counter as C
            diff_counter = C(diffs)
            most_common_diff, freq = diff_counter.most_common(1)[0]

            if freq >= len(diffs) * 0.8 and freq >= 3 and most_common_diff != 0:
                p = self._binomial_prob(freq, len(diffs), 0.1)
                self._add_finding(
                    "ARITHMETIC_SEQUENCE",
                    f"列 '{c}' 呈等差数列: 公差={most_common_diff}，{freq}/{len(diffs)} 个相邻差值一致 ({freq/len(diffs)*100:.1f}%)",
                    p,
                    {
                        "column": c,
                        "common_difference": most_common_diff,
                        "matching_diffs": freq,
                        "total_diffs": len(diffs),
                        "match_rate": freq/len(diffs),
                        "value_range": [min(vals), max(vals)]
                    }
                )

    # ==================================================================
    # 检测 5: 末位数字集中度
    # ==================================================================
    def _detect_digit_preference(self, columns: dict):
        col_names = list(columns.keys())
        for c in col_names:
            vals = columns[c]
            if len(vals) < 5:
                continue

            # 提取末位数字 (保留显著小数位后的末位)
            last_digits = []
            for v in vals:
                s = f"{abs(v):.10g}"
                # 找最后一位数字
                last_char = s.rstrip('0').rstrip('.')
                if last_char:
                    ld = last_char[-1]
                    if ld.isdigit():
                        last_digits.append(int(ld))
                    else:
                        continue
                else:
                    last_digits.append(0)

            if len(last_digits) < 5:
                continue

            digit_counts = Counter(last_digits)

            # 0+5集中度
            zero_five = digit_counts.get(0, 0) + digit_counts.get(5, 0)
            zero_five_pct = zero_five / len(last_digits) * 100

            if zero_five_pct > 50:  # 自然数据约20%
                p = self._binomial_prob(zero_five, len(last_digits), 0.20)
                self._add_finding(
                    "DIGIT_PREFERENCE",
                    f"列 '{c}' 末位数字0+5集中度 = {zero_five_pct:.1f}% (自然≈20%)，共{len(last_digits)}个数字",
                    p,
                    {
                        "column": c,
                        "zero_five_rate": zero_five_pct,
                        "total_digits": len(last_digits),
                        "digit_distribution": dict(digit_counts.most_common())
                    }
                )

            # 全相同末位
            all_same = digit_counts.most_common(1)[0][1]
            if all_same >= len(last_digits) * 0.8 and len(last_digits) >= 5:
                dominant = digit_counts.most_common(1)[0][0]
                p = self._binomial_prob(all_same, len(last_digits), 0.10)
                self._add_finding(
                    "ALL_SAME_LAST_DIGIT",
                    f"列 '{c}' 末位数字高度集中: 数字{dominant}出现 {all_same}/{len(last_digits)} 次 ({all_same/len(last_digits)*100:.1f}%)",
                    p,
                    {
                        "column": c,
                        "dominant_digit": dominant,
                        "dominant_count": all_same,
                        "total_digits": len(last_digits)
                    }
                )

    # ==================================================================
    # 检测 6: 小数点后位数异常 (已知案例特征)
    # ==================================================================
    def _detect_decimal_anomaly(self, columns: dict, rows: list, headers: list):
        for c_idx, c_name in enumerate(columns.keys()):
            vals = columns[c_name]
            if len(vals) < 5:
                continue

            # 统计每个值的小数点后位数
            decimal_lengths = []
            for v in vals:
                s = f"{abs(v):.10g}"
                if '.' in s:
                    fractional = s.split('.')[1].rstrip('0')
                    decimal_lengths.append(len(fractional))
                else:
                    decimal_lengths.append(0)

            if not decimal_lengths:
                continue

            len_counter = Counter(decimal_lengths)
            # 如果存在两类长度: 大部分N位，小部分N+1位且N+1位的末位数字高度相同
            most_common_len = len_counter.most_common(1)[0][0]
            if most_common_len == 0:
                continue

            # 找N+1位的数据点
            n1_indices = [i for i, d in enumerate(decimal_lengths) if d == most_common_len + 1]
            if len(n1_indices) >= 3:
                # 检查这些N+1位数据的末位
                last_digits_n1 = []
                for i in n1_indices:
                    v = vals[i]
                    s = f"{abs(v):.10g}"
                    if '.' in s:
                        frac = s.split('.')[1].rstrip('0')
                        if frac:
                            last_digits_n1.append(int(frac[-1]))
                if last_digits_n1:
                    ld_counter = Counter(last_digits_n1)
                    most_common_ld, ld_count = ld_counter.most_common(1)[0]
                    if ld_count >= len(last_digits_n1) * 0.7 and len(last_digits_n1) >= 3:
                        self._add_finding(
                            "DECIMAL_PRECISION_ANOMALY",
                            f"列 '{c_name}': {len(n1_indices)}个值有{most_common_len+1}位小数(余为{most_common_len}位)，其中{ld_count}/{len(last_digits_n1)}个末位={most_common_ld}",
                            self._binomial_prob(ld_count, len(last_digits_n1), 0.18),
                            {
                                "column": c_name,
                                "common_precision": most_common_len,
                                "anomalous_count": len(n1_indices),
                                "dominant_last_digit": most_common_ld,
                                "dominant_count": ld_count
                            }
                        )

    # ==================================================================
    # 检测 A12: 🆕 v8.4 Decimal前缀重复（来源: 公开方法 numeric_screen）
    # ==================================================================
    def _detect_decimal_prefix_repeat(self, columns: dict):
        """
        检测小数点后前2-3位数字前缀的异常重复模式。
        
        公开方法论：编造数据的人倾向于重复输入相同的小数前缀，
        如 1.234, 1.235, 1.236 的前缀 "23" 重复出现。
        通过统计每列小数值的 decimal_prefix 重复率来检测。
        
        要求：
        - 至少10个含小数点后至少 places 位的数值
        - 重复组>5 或 最高频前缀≥3 或 重复率>15% → 可疑
        """
        for c_name in columns:
            vals = columns[c_name]
            if len(vals) < 10:
                continue
            
            # 转为字符串，提取小数点后前缀
            for places in [2, 3]:
                prefixes = []
                for v in vals:
                    # 处理科学记数法
                    s = f"{abs(v):.10g}"
                    if 'e' in s.lower():
                        mantissa = s.split('e')[0].split('E')[0]
                    else:
                        mantissa = s
                    if '.' not in mantissa:
                        continue
                    frac = mantissa.split('.', 1)[1]
                    # 去掉末尾非数字字符
                    frac = ''.join(c for c in frac if c.isdigit())
                    if len(frac) >= places:
                        prefixes.append(frac[:places])
                
                if len(prefixes) < 10:
                    continue
                
                from collections import Counter as C
                counts = C(prefixes)
                repeated_groups = sum(1 for c in counts.values() if c >= 2)
                repeated_items = sum(c for c in counts.values() if c >= 2)
                max_prefix, max_count = counts.most_common(1)[0]
                repeat_rate = repeated_items / len(prefixes)
                
                flagged = repeated_groups > 5 or max_count >= 3 or repeat_rate > 0.15
                if not flagged:
                    continue
                
                severity = "HIGH" if max_count >= 4 or repeat_rate > 0.25 else "MEDIUM"
                p = self._binomial_prob(
                    repeated_items, len(prefixes),
                    0.10 ** places  # 自然期望: 每个前缀概率 ~10^{-places}
                )
                self._add_finding(
                    "DECIMAL_PREFIX_REPEAT",
                    f"列 '{c_name}' 小数点后{places}位前缀重复: {repeated_groups}组重复, "
                    f"最高频 '{max_prefix}' 出现{max_count}次, 重复率{repeat_rate*100:.1f}%",
                    p,
                    {
                        "column": c_name,
                        "decimal_places": places,
                        "n_values": len(prefixes),
                        "repeated_groups": repeated_groups,
                        "repeated_items": repeated_items,
                        "repeat_rate": round(repeat_rate, 4),
                        "max_prefix": max_prefix,
                        "max_count": max_count,
                        "top_repeats": dict(counts.most_common(10)),
                        "note": "公开方法论: 编造数据常在小数前缀上出现异常重复"
                    }
                )
    
    # ==================================================================
    # 检测 7: 块级重复（多列一起重复）
    # ==================================================================
    def _detect_block_repetition(self, columns: dict):
        col_names = list(columns.keys())
        if len(col_names) < 2:
            return

        # 取最短列长度
        lengths = [len(columns[c]) for c in col_names]
        min_len = min(lengths)
        if min_len < 4:
            return

        # 将行转为元组
        rows_tuples = []
        for i in range(min_len):
            row = tuple(round(columns[c][i], 6) if i < len(columns[c]) else None for c in col_names)
            rows_tuples.append(row)

        tuple_counts = Counter(rows_tuples)
        total_unique = len(tuple_counts)
        total_rows = len(rows_tuples)
        duplicates = total_rows - total_unique

        if duplicates >= 3 and duplicates / total_rows > 0.3:
            self._add_finding(
                "BLOCK_DUPLICATION",
                f"数据块重复: {duplicates}/{total_rows} 行是已存在行的精确副本 ({duplicates/total_rows*100:.1f}%)",
                self._binomial_prob(total_unique, total_rows, 0.5),
                {"total_rows": total_rows, "unique_rows": total_unique, "duplicates": duplicates}
            )

    # ==================================================================
    # 检测 8: 🆕 v5.0 噪声环境下的编造模式检测
    # ==================================================================
    def _detect_noise_pattern(self, columns: dict):
        """
        检测带噪声的编造数据。
        
        当编造者在编造数据中加入随机噪声(±0.2-0.3)后,固定差值检测失效。
        但以下模式仍然可检测:
        1. 底层数字偏好 — 编造者从"好看"的数字开始加噪声
        2. 噪声方差异常 — 正常实验数据噪声方差比编造噪声大
        3. 列间相关性 — 编造的多列从同一底层模板生成
        """
        col_names = list(columns.keys())
        if len(col_names) < 2:
            return

        min_len = min(len(columns[c]) for c in col_names)
        if min_len < 8:
            return

        # 检测1: 噪声列的底层模板检测
        # 如果多列数据与一个隐藏模板列高度相关(r>0.95)
        import math as m
        for c1, c2 in combinations(col_names, 2):
            vals1 = columns[c1][:min_len]
            vals2 = columns[c2][:min_len]
            
            # 计算Pearson相关系数
            mean1 = sum(vals1) / len(vals1)
            mean2 = sum(vals2) / len(vals2)
            
            cov = sum((vals1[i] - mean1) * (vals2[i] - mean2) for i in range(min_len))
            std1 = m.sqrt(sum((v - mean1)**2 for v in vals1))
            std2 = m.sqrt(sum((v - mean2)**2 for v in vals2))
            
            if std1 > 0 and std2 > 0:
                r = cov / (std1 * std2)
                # 高度相关但不是完全相同 (r=1.0 已被固定差值检测覆盖)
                if 0.92 <= r < 0.999:
                    self._add_finding(
                        "NOISE_CORRELATION",
                        f"列 '{c1}' 和 '{c2}' 高度相关 (r={r:.4f})但非精确复制——可能来自同一底层模板加噪声",
                        self._binomial_prob(int(min_len * 0.9), min_len, 0.5),
                        {
                            "column_a": c1, "column_b": c2,
                            "correlation": r,
                            "sample_count": min_len,
                            "note": "噪声环境下的潜在模板编造。需结合末位数字偏好综合判断。"
                        }
                    )

    # ==================================================================
    # 检测 9: 🆕 v5.2 byte-identical 重复检测（公开方法方法论）
    # ==================================================================
    def _detect_byte_identical(self, columns: dict):
        """
        检测同一篇论文中多处数据存在的完全逐字节相同的重复块。
        
        公开方法方法论关键维度: 同一篇论文内多处数据块完全byte-identical，
        在物理上不可能独立产生——这是硬性造假证据。
        
        检测策略:
        1. 行级byte-identical: 每行所有数值转为标准化字符串，检查完全相同的行
        2. 块级byte-identical: 连续k行的序列在论文中不同位置重复出现
        3. 跨列byte-identical: 两列完全逐值相同
        
        与块重复检测(检测7)的区别:
        - 检测7: 数值级别的精确重复
        - 检测9: byte级别的完全一致 + 块级序列重复
        """
        col_names = list(columns.keys())
        if len(col_names) < 2:
            return
        
        min_len = min(len(columns[c]) for c in col_names)
        if min_len < 4:
            return
        
        # 将每行标准化为固定精度的字符串
        row_bytes = []
        for i in range(min_len):
            row_str = "|".join(
                f"{columns[c][i]:.10g}" if i < len(columns[c]) else "NA"
                for c in col_names
            )
            row_bytes.append(row_str)
        
        # === 策略1: 行级byte-identical ===
        byte_counter = Counter(row_bytes)
        total_byte_dupes = sum(c - 1 for c in byte_counter.values() if c > 1)
        
        if total_byte_dupes >= 3:
            duped_groups = []
            for rb, count in byte_counter.items():
                if count >= 3:
                    indices = [i + 1 for i, r in enumerate(row_bytes) if r == rb]
                    duped_groups.append({
                        "row_indices": indices,
                        "count": count,
                        "row_preview": rb[:120]
                    })
            
            if duped_groups:
                p = self._binomial_prob(total_byte_dupes, min_len, 0.01)
                self._add_finding(
                    "BYTE_IDENTICAL_ROWS",
                    f"整行byte-identical重复: {total_byte_dupes}行是其他行的完全字节级副本，分布在{len(duped_groups)}个重复组",
                    p,
                    {
                        "total_rows": min_len,
                        "unique_rows": len(byte_counter),
                        "duplicate_row_count": total_byte_dupes,
                        "duplicate_groups": len(duped_groups),
                        "groups": duped_groups[:5],
                        "note": "公开方法方法: 同一论文内多处数据完全逐字节相同，物理上不可能独立产生"
                    }
                )
        
        # === 策略2: 块级byte-identical ===
        for k in [3, 4, 5]:
            if min_len < k * 2:
                continue
            
            block_positions = {}
            for start in range(min_len - k + 1):
                block = tuple(row_bytes[start:start + k])
                if block not in block_positions:
                    block_positions[block] = []
                block_positions[block].append(start)
            
            for block, positions in block_positions.items():
                if len(positions) >= 2:
                    # 跳过全NA块
                    block_str = "".join(block)
                    if all("NA" in b for b in block):
                        continue
                    pos_str = ", ".join(f"行{p+1}-{p+k}" for p in positions)
                    self._add_finding(
                        "BYTE_IDENTICAL_BLOCK",
                        f"连续{k}行数据块完全byte-identical，出现{len(positions)}次: {pos_str}",
                        self._binomial_prob(len(positions), min_len - k + 1, 0.005),
                        {
                            "block_size": k,
                            "occurrences": len(positions),
                            "positions": [f"行{p+1}-{p+k}" for p in positions],
                            "block_preview": block_str[:200],
                            "note": "块级byte-identical序列重复: 多行精确副本在不同位置独立出现"
                        }
                    )
        
        # === 策略3: 跨列byte-identical ===
        for c1, c2 in combinations(col_names, 2):
            n = min(len(columns[c1]), len(columns[c2]))
            if n < 3:
                continue
            vals1 = [f"{columns[c1][i]:.10g}" for i in range(n)]
            vals2 = [f"{columns[c2][i]:.10g}" for i in range(n)]
            
            if vals1 == vals2:
                self._add_finding(
                    "BYTE_IDENTICAL_COLUMNS",
                    f"列 '{c1}' 和 '{c2}' 全部{n}个值完全逐字节相同——独立实验数据不可能发生",
                    0.0,
                    {
                        "column_a": c1,
                        "column_b": c2,
                        "identical_values": n,
                        "note": "两列独立实验数据完全逐字节相同=硬性编造证据"
                    }
                )
        
        # === 策略4: 🆕 v6.4 列复制改1格（公开方法方法论盲区补完）===
        for c1, c2 in combinations(col_names, 2):
            n = min(len(columns[c1]), len(columns[c2]))
            if n < 4:
                continue
            vals1 = [f"{columns[c1][i]:.10g}" for i in range(n)]
            vals2 = [f"{columns[c2][i]:.10g}" for i in range(n)]
            
            # 计算逐值不同的个数
            diff_indices = []
            for i in range(n):
                if vals1[i] != vals2[i]:
                    diff_indices.append(i)
            
            # 编辑距离=1: 仅1个值不同，其余完全相同
            if len(diff_indices) == 1:
                idx = diff_indices[0]
                self._add_finding(
                    "BYTE_IDENTICAL_NEAR_COLUMNS",
                    f"列 '{c1}' 和 '{c2}' 仅有1个值不同（第{idx+1}行: {vals1[idx]} vs {vals2[idx]}），其余{n-1}行完全逐字节相同——典型的'复制后改1格'造假",
                    0.0,
                    {
                        "column_a": c1,
                        "column_b": c2,
                        "identical_count": n - 1,
                        "total_rows": n,
                        "diff_index": idx + 1,
                        "value_a": vals1[idx],
                        "value_b": vals2[idx],
                        "note": "公开方法方法论盲区: 列复制改1格——两列独立实验数据仅1个值不同=编辑距离1"
                    }
                )
            
            # 编辑距离=2: 仅2个值不同（放宽版本，但需要高相同率）
            elif len(diff_indices) == 2 and n >= 8:
                pct_identical = (n - 2) / n
                if pct_identical >= 0.75:
                    self._add_finding(
                        "BYTE_IDENTICAL_NEAR_COLUMNS",
                        f"列 '{c1}' 和 '{c2}' 仅2个值不同（第{diff_indices[0]+1}行和第{diff_indices[1]+1}行），其余{n-2}/{n}行({pct_identical*100:.0f}%)完全逐字节相同",
                        0.0,
                        {
                            "column_a": c1,
                            "column_b": c2,
                            "identical_count": n - 2,
                            "total_rows": n,
                            "edit_distance": 2,
                            "diff_indices": [i + 1 for i in diff_indices],
                            "note": "公开方法方法论盲区: 列间编辑距离2，高度近似=可疑"
                        }
                    )

    # ==================================================================
    # 检测 13: 🆕 v6.4 等差网格检测（公开方法方法论盲区补完）
    # ==================================================================
    def _detect_arithmetic_grid(self, columns: dict):
        """
        检测取值集合本身呈现等间距分布的"等差网格"模式。
        
        公开方法方法论中的"17个唯一值=等差网格"意指：一列数据的唯一值集合
        本身就是一个等差数列——每个取值是前一个值加固定步长。
        
        与检测4(单列等差数列)的区别:
        - 检测4: 按索引顺序检查相邻差值的重复率
        - 检测13: 检查唯一值集合（忽略出现顺序和频率）是否是等差数列
        
        这是更隐蔽的编造模式——编造者让取值均匀分布在一段区间内，
        看起来"数据分布均匀"，但真正的自然数据不可能呈现精确的等间距。
        """
        for c_name in columns:
            vals = columns[c_name]
            if len(vals) < 5:
                continue
            
            unique_vals = sorted(set(vals))
            n_unique = len(unique_vals)
            
            if n_unique < 4:
                continue
            
            # 计算相邻唯一值的差值
            gaps = [round(unique_vals[i+1] - unique_vals[i], 10) 
                    for i in range(n_unique - 1)]
            
            from collections import Counter as CC
            gap_counter = CC(gaps)
            most_common_gap, gap_freq = gap_counter.most_common(1)[0]
            
            # 如果唯一值占比高+等间距，标记
            unique_ratio = n_unique / len(vals)
            
            # 条件1: 至少一半的唯一值间距完全一致
            # 条件2（v8.1 新增）: 排除常见测量精度导致的天然等间距
            #   0.1/0.5/1.0 是常见精度 → 在连续变量中天然出现
            #   A10 应检测不常见的间距（如 133, 0.25, 0.33）——造假者编造的人为间距
            if (gap_freq >= max(3, n_unique * 0.5) and most_common_gap > 0
                and unique_ratio > 0.5 and n_unique > 10
                and most_common_gap not in (0.1, 0.5, 1.0, 0.01, 0.05)):
                # 条件3: 唯一值占比高说明"每个值恰好出现几次+每个值恰好间隔相同"
                # 这在实际实验中极不可能
                p = self._binomial_prob(gap_freq, n_unique - 1, 0.1)
                self._add_finding(
                    "ARITHMETIC_GRID",
                    f"列 '{c_name}': 唯一值集合呈等差网格分布——{n_unique}个唯一值中{n_unique-1}个相邻间距，"
                    f"有{gap_freq}个间距精确={most_common_gap}（{gap_freq}/{n_unique-1}={gap_freq/(n_unique-1)*100:.0f}%），"
                    f"唯一值占比{unique_ratio*100:.0f}%",
                    p,
                    {
                        "column": c_name,
                        "unique_count": n_unique,
                        "total_count": len(vals),
                        "unique_ratio": unique_ratio,
                        "most_common_gap": most_common_gap,
                        "matching_gaps": gap_freq,
                        "total_gaps": n_unique - 1,
                        "sample_unique_values": unique_vals[:10],
                        "note": "公开方法方法论盲区: 取值集合本身呈等间距分布——自然数据不可能如此整齐"
                    }
                )


    # ==================================================================
    # 检测 10: 🆕 v6.0 常数平移检测（已知公开案例）
    # ==================================================================
    def _detect_constant_shift(self, columns: dict):
        """
        检测一个序列是另一个序列的常数平移。
        
        与检测1(固定差值)的区别:
        - 检测1: 需要大量行(>=3且>=50%行差值相同)才触发
        - 检测10: 检测整个序列的每个对应位置差值是否完全相同（包括只有3个值的情况）
        
        已知公开案例:
        GlueBody2 [1.03, 0.99, 0.97]
        D3       [1.05, 1.01, 0.99]
        每个对应位差值 = +0.02
        
        造假者担心两组数据完全一样太明显，加一个常数来"差异化"——
        但这正是"一个序列=另一个序列+常数"的铁证。
        """
        col_names = list(columns.keys())
        for c1, c2 in combinations(col_names, 2):
            vals1 = columns[c1]
            vals2 = columns[c2]
            n = min(len(vals1), len(vals2))
            if n < 3:
                continue
            
            # 计算对应位差值
            diffs = [round(vals2[i] - vals1[i], 10) for i in range(n)]
            # 检查所有差值是否相同（允许浮点舍入误差）
            unique_diffs = set(diffs)
            
            if len(unique_diffs) == 1 and diffs[0] != 0:
                shift_value = diffs[0]
                # 计算自然概率: n个独立差值都恰好相同的概率极低
                # 保守假设每个差值有10个可能值（0-9末位），相同概率=10%
                p = self._binomial_prob(n, n, 0.1)
                
                self._add_finding(
                    "CONSTANT_SHIFT",
                    f"列 '{c2}' 的每个值 = 列 '{c1}' 的对应值 + {shift_value}，全部{n}个对应位置差值完全相同",
                    p,
                    {
                        "column_a": c1, "column_b": c2,
                        "shift_value": shift_value,
                        "total_pairs": n,
                        "sample_values_1": vals1[:5],
                        "sample_values_2": vals2[:5],
                        "note": "一个序列是另一个序列的常数平移——独立实验数据不可能呈现这种规律"
                    }
                )

    # ==================================================================
    # 检测 11: 🆕 v6.0 跨组行级复制粘贴（已知公开案例）
    # ==================================================================
    def _detect_cross_group_row_dup(self, columns: dict):
        """
        检测两组声称独立的实验数据中，前N行完全相同但后续不同。
        
        与检测7(块级重复)的区别:
        - 检测7: 检查整表内所有行tuple是否有重复，要求重复率>30%才触发
        - 检测11: 专门检查"两组的前N行完全相同"——这是典型的复制粘贴造假模式
          编造者复制了一组数据到另一组，只改了最后一个值来"差异化"
        
        已知公开案例:
        GlueBody  前三个样本: [14.5, 16.0, 21.5]
        GlueBody2 前三个样本: [14.5, 16.0, 21.5]
        只有第四个样本不同。
        
        检测方法: 对任意两列，检查前k行(3≤k≤total-1)完全相同
        """
        col_names = list(columns.keys())
        for c1, c2 in combinations(col_names, 2):
            vals1 = columns[c1]
            vals2 = columns[c2]
            n = min(len(vals1), len(vals2))
            if n < 4:
                continue
            
            # 找第一个不同的位置
            first_diff_idx = None
            for i in range(n):
                if abs(vals1[i] - vals2[i]) > 1e-8:
                    first_diff_idx = i
                    break
            
            # 如果没有不同→就是BYTE_IDENTICAL_COLUMNS，已由检测9处理
            if first_diff_idx is None:
                continue
            
            # v8.1 修正阈值: 前k个完全相同，k>=5且k/n>=0.15 (n=60时至少9行)
            # 原阈值 k>=3 在随机数据中FPR≈0.5%（200次中触发1次）
            # 随机正态分布中前5行恰好完全相同的概率远低于前3行
            if first_diff_idx >= 5 and first_diff_idx / n >= 0.15:
                k = first_diff_idx
                # 验证前k个完全相同
                all_same = all(
                    abs(vals1[i] - vals2[i]) < 1e-8 for i in range(k)
                )
                if all_same and n >= 4:
                    p = self._binomial_prob(k, n, 0.1)
                    self._add_finding(
                        "CROSS_GROUP_ROW_DUPLICATION",
                        f"列 '{c1}' 和 '{c2}' 的前{k}行完全相同（{k}/{n}={k/n*100:.0f}%），仅第{k+1}行起不同——独立实验组不可能前{k}行完全一样",
                        p,
                        {
                            "column_a": c1, "column_b": c2,
                            "identical_prefix_length": k,
                            "total_rows": n,
                            "identical_preview": vals1[:k],
                            "first_difference_at": {
                                "row": k + 1,
                                "value_a": vals1[k],
                                "value_b": vals2[k]
                            },
                            "note": "跨组行级复制粘贴: 编造者复制一组数据到另一组，只改了最后的值"
                        }
                    )

    # ==================================================================
    # 检测 12: 🆕 v6.0 跨文件/跨sheet固定差值配对
    # ==================================================================
    def detect_cross_file(self, filepaths: list) -> list:
        """
        从多个文件加载所有数值列，进行跨文件固定差值配对检测。
        
        区别于单文件内检测1: 造假者可能把数据分散到不同文件/sheet中，
        单文件检测无法发现跨文件的固定差值规律。
        
        已知公开案例:
        图3G:     [33.8, 33.9, 37.9]
        扩展图7B: [75.2, 75.3, 79.3]
        每个对应位差值 = 41.4
        
        Args:
            filepaths: 文件路径列表 (CSV或Excel)
        Returns:
            与detect()相同格式的findings列表，额外标注了跨文件信息
        """
        self.multi_source_mode = True
        self.findings = []
        
        # 加载所有文件的所有数值列
        for fp in filepaths:
            source_name = Path(fp).stem
            try:
                columns = self._load_single_file(fp)
                self.all_source_columns[source_name] = columns
            except Exception as e:
                self.findings.append({
                    "error": f"无法加载 {fp}: {e}",
                    "source": source_name
                })
        
        # 跨文件配对检测
        self._detect_cross_file_pairs()
        
        # 也做单文件内部检测（多层防护）
        for source_name, columns in self.all_source_columns.items():
            self._detect_arithmetic_grid(columns)
            self._detect_constant_shift(columns)
            self._detect_cross_group_row_dup(columns)
        
        return self.findings
    
    def _load_single_file(self, filepath: str) -> dict:
        """加载单个文件的所有数值列"""
        path = Path(filepath)
        if path.suffix.lower() in ('.xlsx', '.xls'):
            return self._load_excel_columns(filepath)
        else:
            return self._load_csv_columns(filepath)
    
    def _load_csv_columns(self, filepath: str) -> dict:
        """从CSV加载数值列"""
        rows = []
        with open(filepath, 'r', encoding='utf-8-sig') as f:
            reader = csv.reader(f)
            headers = next(reader, None)
            for row in reader:
                rows.append(row)
        return self._build_columns(headers, rows)
    
    def _load_excel_columns(self, filepath: str) -> dict:
        """从Excel加载所有sheet的数值列，sheet名作为前缀"""
        if not HAS_OPENPYXL:
            return {}
        all_cols = {}
        wb = openpyxl.load_workbook(filepath, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = [[cell.value for cell in row] for row in ws.iter_rows()]
            if not rows:
                continue
            headers = rows[0]
            data_rows = rows[1:]
            cols = self._build_columns(headers, data_rows)
            for col_name, values in cols.items():
                all_cols[f"{sheet_name}/{col_name}"] = values
        return all_cols
    
    def _build_columns(self, headers: list, rows: list) -> dict:
        """从表头和数据行构建数值列"""
        columns = {}
        for i, h in enumerate(headers):
            col_data = []
            for row in rows:
                if i < len(row):
                    val = self._parse_number(row[i])
                    if val is not None:
                        col_data.append(val)
            if len(col_data) >= 3:
                name = str(h).strip() if h else f"Col_{i}"
                columns[name] = col_data
        return columns
    
    def _detect_cross_file_pairs(self):
        """
        跨文件/跨sheet固定差值配对检测。
        
        从不同来源提取的所有数值列，两两配对检测固定差值。
        这是已知公开案例的核心检测维度。
        """
        # 构建全局列列表: [(source_name, col_name, values)]
        all_cols = []
        for source_name, columns in self.all_source_columns.items():
            for col_name, values in columns.items():
                all_cols.append((source_name, col_name, values))
        
        if len(all_cols) < 2:
            return
        
        for (s1, n1, v1), (s2, n2, v2) in combinations(all_cols, 2):
            total = min(len(v1), len(v2))
            if total < 3:
                continue
            
            diffs = [round(v2[i] - v1[i], 10) for i in range(total)]
            unique_diffs = set(diffs)
            
            # 检测固定差值: 所有对应位差值相同
            if len(unique_diffs) == 1 and diffs[0] != 0:
                shift = diffs[0]
                p = self._binomial_prob(total, total, 0.1)
                self._add_finding(
                    "CROSS_FILE_FIXED_DIFFERENCE",
                    f"跨文件固定差值: [{s1}] {n1} vs [{s2}] {n2}，"
                    f"全部{total}个对应位置差值精确={shift}",
                    p,
                    {
                        "source_a": s1, "column_a": n1,
                        "source_b": s2, "column_b": n2,
                        "fixed_difference": shift,
                        "total_pairs": total,
                        "sample_a": v1[:5],
                        "sample_b": v2[:5],
                        "note": "不同文件/图表的数据呈现完全相同差值的对应关系——"
                                "独立实验不可能出现"
                    }
                )
            
            # 检测固定差值集中度（大部分但不一定全部相同）
            elif total >= 5:
                from collections import Counter as CC
                diff_counter = CC(diffs)
                most_common_diff, freq = diff_counter.most_common(1)[0]
                if freq >= total * 0.8 and freq >= 4:
                    p = self._binomial_prob(freq, total, 0.1)
                    self._add_finding(
                        "CROSS_FILE_FIXED_DIFFERENCE",
                        f"跨文件固定差值: [{s1}] {n1} vs [{s2}] {n2}，"
                        f"差值={most_common_diff} 出现{freq}/{total}次 ({freq/total*100:.0f}%)",
                        p,
                        {
                            "source_a": s1, "column_a": n1,
                            "source_b": s2, "column_b": n2,
                            "fixed_difference": most_common_diff,
                            "matching_pairs": freq,
                            "total_pairs": total,
                            "match_rate": freq/total,
                            "note": "跨文件数据呈现高度集中的差值——高度人工特征"
                        }
                    )

    # ==================================================================
    # 辅助函数
    # ==================================================================
    @staticmethod
    def _parse_number(val):
        if val is None:
            return None
        if isinstance(val, (int, float)):
            if math.isnan(val) or math.isinf(val):
                return None
            return val
        if isinstance(val, str):
            val = val.strip().replace('%', '').replace(',', '')
            try:
                return float(val)
            except ValueError:
                return None
        return None

    @staticmethod
    def _binomial_prob(k: int, n: int, p: float) -> float:
        """计算二项分布的上尾概率 P(X >= k | n, p) 近似"""
        if k == 0 or n == 0:
            return 1.0
        import math as m
        mean = n * p
        std = m.sqrt(n * p * (1 - p))
        if std == 0:
            return 0.0 if k > mean else 1.0
        z = (k - mean) / std
        if z > 20:
            return 0.0
        if z < 0:
            return 1.0
        # 正态近似
        return 0.5 * m.erfc(z / m.sqrt(2))

    def _add_finding(self, detection_type: str, description: str, p_value: float, details: dict):
        self.findings.append({
            "type": detection_type,
            "description": description,
            "p_value": f"{p_value:.2e}" if p_value > 1e-300 else "≈ 0",
            "impossible_probability": "p < 10^-18" if p_value < 1e-18 else (
                f"p = {p_value:.2e}" if p_value < 0.001 else f"p = {p_value:.4f}"
            ),
            "severity": "CRITICAL" if p_value < 1e-10 else (
                "HIGH" if p_value < 1e-5 else "MEDIUM"
            ),
            "details": details
        })


def main():
    import argparse
    parser = argparse.ArgumentParser(description='等差数列/固定比值自动检测器 v6.0')
    parser.add_argument('input', nargs='*', help='CSV或Excel文件路径（可多个，跨文件检测）')
    parser.add_argument('-j', '--json', action='store_true', help='JSON格式输出')
    parser.add_argument('-x', '--cross-file', action='store_true',
                        help='跨文件模式: 从多个文件加载所有数值列做跨文件配对检测')
    args = parser.parse_args()

    detector = ArithmeticSequenceDetector()
    
    if args.cross_file and len(args.input) >= 2:
        findings = detector.detect_cross_file(args.input)
        output = {
            "mode": "cross_file",
            "input_files": args.input,
            "sources_count": len(detector.all_source_columns),
            "findings_count": len(findings),
            "findings": findings
        }
    elif len(args.input) == 1:
        findings = detector.detect_from_file(args.input[0])
        output = {
            "input_file": args.input[0],
            "total_columns": detector.total_columns,
            "total_cells": detector.total_cells,
            "findings_count": len(findings),
            "findings": findings
        }
    elif len(args.input) > 1:
        # 多个文件但未指定跨文件标志 → 逐个检测
        all_findings = []
        for fp in args.input:
            findings = detector.detect_from_file(fp)
            all_findings.extend(findings)
        output = {
            "input_files": args.input,
            "total_columns": detector.total_columns,
            "total_cells": detector.total_cells,
            "findings_count": len(all_findings),
            "findings": all_findings
        }
    else:
        parser.print_help()
        return

    if args.json:
        print(json.dumps(output, ensure_ascii=False, indent=2))
    else:
        from pathlib import Path
        if args.cross_file:
            print(f"=== 跨文件算术异常检测报告 ===")
            print(f"扫描: {len(args.input)} 个文件, {len(detector.all_source_columns)} 个来源")
        elif len(args.input) == 1:
            print(f"=== 算术异常检测报告: {args.input[0]} ===")
            print(f"扫描: {detector.total_columns} 列, {detector.total_cells} 个数据点")
        else:
            print(f"=== 批量算术异常检测报告: {len(args.input)} 个文件 ===")
        print(f"发现: {len(output['findings'])} 个异常模式\n")
        for i, f in enumerate(output['findings'], 1):
            if 'error' in f:
                print(f"  [ERROR] {f['error']}")
                continue
            print(f"  [{f.get('severity', 'INFO')}] #{i} {f['type']}")
            print(f"  {f['description']}")
            print(f"  概率: {f.get('impossible_probability', 'N/A')}")
            print()

        if not output['findings']:
            print("  ✅ 未发现明显的数学规律异常。")

    # 输出JSON供管道使用
    if not args.json:
        print("\n--- JSON OUTPUT ---")
        print(json.dumps(output, ensure_ascii=False, indent=2))


if __name__ == '__main__':
    main()
